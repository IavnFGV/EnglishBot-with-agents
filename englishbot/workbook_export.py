from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook

from .assets import PRIMARY_AUDIO_ROLE, PRIMARY_IMAGE_ROLE
from .bulk_edit import get_bulk_edit_export_dir
from .config import get_infra_static_base_url
from .db import get_connection, utc_now

WORKBOOK_VERSION = "family_bulk_edit_v1"
META_SHEET = "meta"
LEARNING_ITEMS_SHEET = "learning_items"
LEARNING_ITEM_COLUMNS = (
    "item_key",
    "text",
    "translation_ru",
    "translation_uk",
    "translation_bg",
    "topics",
    "image_ref",
    "image",
    "audio_ref",
    "is_archived",
)


@dataclass(frozen=True)
class FamilyWorkbookExportResult:
    file_path: Path
    topic_count: int
    learning_item_count: int


def export_family_workbook(family_id: int, *, output_path: Path | None = None) -> FamilyWorkbookExportResult:
    if output_path is None:
        output_path = get_bulk_edit_export_dir() / f"family-{family_id}__{utc_now().replace(':', '-')}.xlsx"

    static_base_url = get_infra_static_base_url()
    workbook = Workbook()
    meta_sheet = workbook.active
    meta_sheet.title = META_SHEET
    meta_sheet.append(("field", "value"))
    meta_sheet.append(("version", WORKBOOK_VERSION))
    meta_sheet.append(("family_id", family_id))
    meta_sheet.append(("exported_at_utc", utc_now()))

    learning_items_sheet = workbook.create_sheet(LEARNING_ITEMS_SHEET)
    learning_items_sheet.append(LEARNING_ITEM_COLUMNS)

    topic_titles_by_item_id, topic_count = _load_topic_titles_by_item_id(family_id)
    for row in _list_family_learning_item_rows(family_id):
        image_ref = _build_export_image_ref(
            image_ref=str(row["image_ref"] or ""),
            image_source_url=str(row["image_source_url"] or ""),
            static_base_url=static_base_url,
        )
        image_formula = _build_image_formula(
            image_ref=image_ref,
            row_number=learning_items_sheet.max_row + 1,
        )
        learning_items_sheet.append(
            (
                f"item-{int(row['id'])}",
                str(row["text"]),
                str(row["translation_ru"] or ""),
                str(row["translation_uk"] or ""),
                str(row["translation_bg"] or ""),
                "\n".join(topic_titles_by_item_id.get(int(row["id"]), [])),
                image_ref,
                image_formula,
                str(row["audio_ref"] or ""),
                int(row["is_archived"]),
            )
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return FamilyWorkbookExportResult(
        file_path=output_path,
        topic_count=topic_count,
        learning_item_count=len(_list_family_learning_item_rows(family_id)),
    )


def _list_family_learning_item_rows(family_id: int) -> list[dict[str, object]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT
                learning_items.id,
                learning_items.text,
                learning_items.is_archived,
                MAX(CASE WHEN translations.language_code = 'ru' THEN translations.translation_text END) AS translation_ru,
                MAX(CASE WHEN translations.language_code = 'uk' THEN translations.translation_text END) AS translation_uk,
                MAX(CASE WHEN translations.language_code = 'bg' THEN translations.translation_text END) AS translation_bg,
                image_assets.local_path AS image_local_path,
                image_assets.source_url AS image_source_url,
                audio_assets.local_path AS audio_local_path,
                audio_assets.source_url AS audio_source_url
            FROM learning_items
            LEFT JOIN learning_item_translations AS translations
              ON translations.learning_item_id = learning_items.id
            LEFT JOIN learning_item_assets AS image_links
              ON image_links.learning_item_id = learning_items.id
             AND image_links.role = ?
            LEFT JOIN assets AS image_assets
              ON image_assets.id = image_links.asset_id
            LEFT JOIN learning_item_assets AS audio_links
              ON audio_links.learning_item_id = learning_items.id
             AND audio_links.role = ?
            LEFT JOIN assets AS audio_assets
              ON audio_assets.id = audio_links.asset_id
            WHERE learning_items.family_id = ?
            GROUP BY
                learning_items.id,
                learning_items.text,
                learning_items.is_archived,
                image_assets.local_path,
                image_assets.source_url,
                audio_assets.local_path,
                audio_assets.source_url
            ORDER BY learning_items.id
            """,
            (
                PRIMARY_IMAGE_ROLE,
                PRIMARY_AUDIO_ROLE,
                family_id,
            ),
        ).fetchall()
    return [
        {
            "id": int(row["id"]),
            "text": str(row["text"]),
            "is_archived": int(row["is_archived"]),
            "translation_ru": row["translation_ru"],
            "translation_uk": row["translation_uk"],
            "translation_bg": row["translation_bg"],
            "image_ref": row["image_local_path"] or row["image_source_url"],
            "image_source_url": row["image_source_url"],
            "audio_ref": row["audio_local_path"] or row["audio_source_url"],
        }
        for row in rows
    ]


def _load_topic_titles_by_item_id(family_id: int) -> tuple[dict[int, list[str]], int]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT
                topic_items.learning_item_id,
                topics.title
            FROM topic_items
            JOIN topics
              ON topics.id = topic_items.topic_id
            JOIN learning_items
              ON learning_items.id = topic_items.learning_item_id
            WHERE topics.family_id = ?
              AND learning_items.family_id = ?
            ORDER BY topics.title, topic_items.id
            """,
            (family_id, family_id),
        ).fetchall()
        topic_count = int(
            connection.execute(
                """
                SELECT COUNT(*)
                FROM topics
                WHERE family_id = ?
                """,
                (family_id,),
            ).fetchone()[0]
        )
    topic_titles_by_item_id: dict[int, list[str]] = {}
    for row in rows:
        topic_titles_by_item_id.setdefault(int(row["learning_item_id"]), []).append(str(row["title"]))
    return topic_titles_by_item_id, topic_count


def _build_export_image_ref(*, image_ref: str, image_source_url: str, static_base_url: str | None) -> str:
    return (
        _build_public_image_url(image_ref=image_ref, static_base_url=static_base_url)
        or image_source_url.strip()
        or image_ref.strip()
    )


def _build_image_formula(*, image_ref: str, row_number: int) -> str:
    if not image_ref.startswith(("http://", "https://")):
        return ""
    return f"=IMAGE(G{row_number})"


def _build_public_image_url(*, image_ref: str, static_base_url: str | None) -> str | None:
    if static_base_url is None:
        return None

    normalized_ref = image_ref.strip().replace("\\", "/")
    if not normalized_ref or normalized_ref.startswith(("http://", "https://")):
        return None

    if normalized_ref.startswith("/app/assets/"):
        public_path = normalized_ref.removeprefix("/app/assets/")
    elif normalized_ref.startswith("assets/"):
        public_path = normalized_ref.removeprefix("assets/")
    else:
        return None

    public_path = public_path.strip("/")
    if not public_path:
        return None
    return f"{static_base_url}/{quote(public_path, safe='/')}"
