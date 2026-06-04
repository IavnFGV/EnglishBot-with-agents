from __future__ import annotations

import re
import sqlite3
from inspect import signature
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from .assets import (
    ASSET_TYPE_AUDIO,
    ASSET_TYPE_IMAGE,
    PRIMARY_AUDIO_ROLE,
    PRIMARY_IMAGE_ROLE,
    download_remote_asset_content,
    finalize_workbook_import_asset,
    replace_learning_item_assets_for_role,
    resolve_runtime_asset_path,
    store_workbook_import_asset,
)
from .bulk_edit import create_bulk_edit_backup
from .db import get_connection, utc_now
from .workbook_export import LEARNING_ITEMS_SHEET, LEARNING_ITEM_COLUMNS, META_SHEET

SUPPORTED_TRANSLATION_COLUMNS = ("translation_ru", "translation_uk", "translation_bg")
TOPIC_NAME_PATTERN = re.compile(r"[^a-z0-9]+")


class WorkbookImportValidationError(ValueError):
    def __init__(self, errors: list[str]) -> None:
        super().__init__("\n".join(errors))
        self.errors = errors


@dataclass(frozen=True)
class WorkbookImportRow:
    row_number: int
    item_key: str
    text: str
    translations: dict[str, str]
    topic_titles: list[str]
    image_ref: str
    audio_ref: str
    is_archived: bool


@dataclass(frozen=True)
class PreparedAssetRef:
    source_url: str | None
    local_path: str
    staged_local_path: str | None = None


@dataclass(frozen=True)
class PreparedWorkbookImportRow:
    row_number: int
    item_key: str
    text: str
    translations: dict[str, str]
    topic_titles: list[str]
    image_asset: PreparedAssetRef | None
    audio_asset: PreparedAssetRef | None
    is_archived: bool


@dataclass(frozen=True)
class PreparedFamilyWorkbookImport:
    rows: list[PreparedWorkbookImportRow]
    staged_asset_paths: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookImportProgress:
    phase: str
    processed_rows: int
    total_rows: int
    current_item_text: str = ""


@dataclass(frozen=True)
class FamilyWorkbookImportSummary:
    created: int
    updated: int
    archived: int
    unchanged: int
    backup_file_path: Path


ProgressCallback = Callable[[WorkbookImportProgress], None]


def validate_family_workbook_import(file_path: Path, family_id: int) -> list[WorkbookImportRow]:
    workbook = load_workbook(file_path, data_only=False)
    if META_SHEET not in workbook.sheetnames:
        raise WorkbookImportValidationError([f"Missing required sheet: {META_SHEET}."])
    if LEARNING_ITEMS_SHEET not in workbook.sheetnames:
        raise WorkbookImportValidationError([f"Missing required sheet: {LEARNING_ITEMS_SHEET}."])

    worksheet = workbook[LEARNING_ITEMS_SHEET]
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    header = [str(value or "").strip() for value in (header_cells or ())]
    if tuple(header) != LEARNING_ITEM_COLUMNS:
        raise WorkbookImportValidationError(
            [f"Unexpected workbook columns. Expected: {', '.join(LEARNING_ITEM_COLUMNS)}."]
        )

    existing_item_ids = _load_existing_family_item_ids(family_id)
    errors: list[str] = []
    rows: list[WorkbookImportRow] = []
    seen_item_keys: set[str] = set()

    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        normalized = {
            column_name: _normalize_cell(value)
            for column_name, value in zip(LEARNING_ITEM_COLUMNS, values, strict=False)
        }
        if not any(normalized.values()):
            continue
        item_key = normalized["item_key"]
        text = normalized["text"]
        if not text:
            errors.append(f"Row {row_number}: text is required.")
        if item_key:
            if item_key in seen_item_keys:
                errors.append(f"Row {row_number}: duplicate item_key {item_key}.")
            seen_item_keys.add(item_key)
            matched_item_id = _parse_item_key(item_key)
            if matched_item_id is None:
                errors.append(f"Row {row_number}: invalid item_key {item_key}.")
            elif matched_item_id not in existing_item_ids:
                errors.append(f"Row {row_number}: item_key {item_key} does not belong to this family.")
        try:
            is_archived = _parse_boolish(normalized["is_archived"])
        except ValueError:
            errors.append(f"Row {row_number}: invalid is_archived value {normalized['is_archived']!r}.")
            is_archived = False

        rows.append(
            WorkbookImportRow(
                row_number=row_number,
                item_key=item_key,
                text=text,
                translations={
                    language_code: normalized[f"translation_{language_code}"]
                    for language_code in ("ru", "uk", "bg")
                    if normalized[f"translation_{language_code}"]
                },
                topic_titles=_parse_topic_titles(normalized["topics"]),
                image_ref=normalized["image_ref"],
                audio_ref=normalized["audio_ref"],
                is_archived=is_archived,
            )
        )

    if errors:
        raise WorkbookImportValidationError(errors)
    return rows


def prepare_family_workbook_import(
    file_path: Path,
    family_id: int,
    rows: list[WorkbookImportRow] | None = None,
    progress_callback: ProgressCallback | None = None,
) -> PreparedFamilyWorkbookImport:
    if rows is None:
        rows = validate_family_workbook_import(file_path, family_id)

    existing_asset_refs = _load_existing_asset_refs(family_id)
    prepared_rows: list[PreparedWorkbookImportRow] = []
    staged_asset_paths: list[str] = []
    total_rows = len(rows)

    try:
        for row_index, row in enumerate(rows, start=1):
            prepared_row = _prepare_row(
                row,
                existing_asset_refs=existing_asset_refs.get(_parse_item_key(row.item_key) or -1, {}),
            )
            prepared_rows.append(prepared_row)
            for asset in (prepared_row.image_asset, prepared_row.audio_asset):
                if asset is not None and asset.staged_local_path:
                    staged_asset_paths.append(asset.staged_local_path)
            _report_progress(
                progress_callback,
                WorkbookImportProgress(
                    phase="preparing",
                    processed_rows=row_index,
                    total_rows=total_rows,
                    current_item_text=row.text,
                ),
            )
    except Exception:
        _cleanup_staged_asset_paths(staged_asset_paths)
        raise

    return PreparedFamilyWorkbookImport(
        rows=prepared_rows,
        staged_asset_paths=tuple(staged_asset_paths),
    )


def cleanup_prepared_family_workbook_import(prepared_import: PreparedFamilyWorkbookImport) -> None:
    _cleanup_staged_asset_paths(prepared_import.staged_asset_paths)


def apply_prepared_family_workbook_import(
    family_id: int,
    prepared_import: PreparedFamilyWorkbookImport,
    backup_file_path: Path,
    progress_callback: ProgressCallback | None = None,
) -> FamilyWorkbookImportSummary:
    total_rows = len(prepared_import.rows)
    summary = {
        "created": 0,
        "updated": 0,
        "archived": 0,
        "unchanged": 0,
    }
    finalized_asset_paths: list[str] = []
    with get_connection() as connection:
        existing_items = _load_existing_items(connection, family_id)
        existing_topics = _load_existing_topics(connection, family_id)
        imported_item_ids: set[int] = set()
        active_topic_membership: dict[str, list[int]] = {}

        try:
            connection.execute("BEGIN")
            for row_index, row in enumerate(prepared_import.rows, start=1):
                imported_item_id, change_kind = _upsert_learning_item(
                    connection,
                    family_id,
                    row,
                    existing_items,
                    finalized_asset_paths=finalized_asset_paths,
                )
                imported_item_ids.add(imported_item_id)
                summary[change_kind] += 1
                if not row.is_archived:
                    for topic_title in row.topic_titles:
                        active_topic_membership.setdefault(topic_title, []).append(imported_item_id)
                _report_progress(
                    progress_callback,
                    WorkbookImportProgress(
                        phase="applying",
                        processed_rows=row_index,
                        total_rows=total_rows,
                        current_item_text=row.text,
                    ),
                )

            for item_id, item in existing_items.items():
                if item_id in imported_item_ids:
                    continue
                if int(item["is_archived"]) == 0:
                    _archive_learning_item(connection, item_id)
                    summary["archived"] += 1

            _sync_topics(connection, family_id, existing_topics, active_topic_membership)
            connection.commit()
        except Exception:
            connection.rollback()
            _cleanup_staged_asset_paths(
                [*prepared_import.staged_asset_paths, *finalized_asset_paths]
            )
            raise

    return FamilyWorkbookImportSummary(
        created=int(summary["created"]),
        updated=int(summary["updated"]),
        archived=int(summary["archived"]),
        unchanged=int(summary["unchanged"]),
        backup_file_path=backup_file_path,
    )


def apply_family_workbook_import(
    file_path: Path,
    family_id: int,
    started_by_user_id: int,
    rows: list[WorkbookImportRow] | None = None,
    progress_callback: ProgressCallback | None = None,
) -> FamilyWorkbookImportSummary:
    if rows is None:
        rows = validate_family_workbook_import(file_path, family_id)
    backup_file_path = create_bulk_edit_backup(family_id=family_id, user_id=started_by_user_id)
    prepared_import = prepare_family_workbook_import(
        file_path,
        family_id,
        rows,
        progress_callback=progress_callback,
    )
    try:
        return apply_prepared_family_workbook_import(
            family_id,
            prepared_import,
            backup_file_path,
            progress_callback=progress_callback,
        )
    except Exception:
        cleanup_prepared_family_workbook_import(prepared_import)
        raise


def _load_existing_family_item_ids(family_id: int) -> set[int]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT id
            FROM learning_items
            WHERE family_id = ?
            """,
            (family_id,),
        ).fetchall()
    return {int(row["id"]) for row in rows}


def _load_existing_asset_refs(
    family_id: int,
) -> dict[int, dict[str, tuple[str, str]]]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT
                learning_items.id AS learning_item_id,
                learning_item_assets.role,
                assets.local_path,
                assets.source_url
            FROM learning_items
            LEFT JOIN learning_item_assets
              ON learning_item_assets.learning_item_id = learning_items.id
            LEFT JOIN assets
              ON assets.id = learning_item_assets.asset_id
            WHERE learning_items.family_id = ?
            """,
            (family_id,),
        ).fetchall()
    result: dict[int, dict[str, tuple[str, str]]] = {}
    for row in rows:
        role = str(row["role"] or "")
        if not role:
            continue
        result.setdefault(int(row["learning_item_id"]), {})[role] = (
            str(row["local_path"] or ""),
            str(row["source_url"] or ""),
        )
    return result


def _load_existing_items(connection: sqlite3.Connection, family_id: int) -> dict[int, sqlite3.Row]:
    rows = connection.execute(
        """
        SELECT id, family_id, lexeme_id, text, is_archived, created_at, updated_at
        FROM learning_items
        WHERE family_id = ?
        ORDER BY id
        """,
        (family_id,),
    ).fetchall()
    return {int(row["id"]): row for row in rows}


def _load_existing_topics(connection: sqlite3.Connection, family_id: int) -> dict[str, sqlite3.Row]:
    rows = connection.execute(
        """
        SELECT id, family_id, name, title, is_archived, created_at, updated_at
        FROM topics
        WHERE family_id = ?
        ORDER BY id
        """,
        (family_id,),
    ).fetchall()
    return {_normalize_topic_identity(str(row["title"])): row for row in rows}


def _prepare_row(
    row: WorkbookImportRow,
    *,
    existing_asset_refs: dict[str, tuple[str, str]],
) -> PreparedWorkbookImportRow:
    return PreparedWorkbookImportRow(
        row_number=row.row_number,
        item_key=row.item_key,
        text=row.text,
        translations=row.translations,
        topic_titles=row.topic_titles,
        image_asset=_prepare_asset_ref(
            row.image_ref,
            role=PRIMARY_IMAGE_ROLE,
            asset_type=ASSET_TYPE_IMAGE,
            existing_ref=existing_asset_refs.get(PRIMARY_IMAGE_ROLE),
        ),
        audio_asset=_prepare_asset_ref(
            row.audio_ref,
            role=PRIMARY_AUDIO_ROLE,
            asset_type=ASSET_TYPE_AUDIO,
            existing_ref=existing_asset_refs.get(PRIMARY_AUDIO_ROLE),
        ),
        is_archived=row.is_archived,
    )


def _prepare_asset_ref(
    asset_ref: str,
    *,
    role: str,
    asset_type: str,
    existing_ref: tuple[str, str] | None,
) -> PreparedAssetRef | None:
    del role
    normalized_ref = asset_ref.strip()
    if not normalized_ref:
        return None
    if not normalized_ref.startswith(("http://", "https://")):
        return PreparedAssetRef(source_url=None, local_path=normalized_ref)

    existing_local_path, existing_source_url = existing_ref or ("", "")
    if existing_source_url == normalized_ref and existing_local_path:
        return PreparedAssetRef(
            source_url=normalized_ref,
            local_path=existing_local_path,
        )

    content = download_remote_asset_content(asset_type, normalized_ref)
    local_path = store_workbook_import_asset(
        asset_type,
        content,
        source_url=normalized_ref,
    )
    return PreparedAssetRef(
        source_url=normalized_ref,
        local_path=local_path,
        staged_local_path=local_path,
    )


def _upsert_learning_item(
    connection: sqlite3.Connection,
    family_id: int,
    row: PreparedWorkbookImportRow,
    existing_items: dict[int, sqlite3.Row],
    *,
    finalized_asset_paths: list[str],
) -> tuple[int, str]:
    existing_item = existing_items.get(_parse_item_key(row.item_key) or -1)
    if existing_item is None:
        item_id = _create_learning_item(connection, family_id, row)
        _sync_translations(connection, item_id, row.translations)
        _sync_media_assets(
            connection,
            item_id,
            row,
            finalized_asset_paths=finalized_asset_paths,
        )
        return item_id, "created"

    lexeme_id = _get_or_create_lexeme_id(connection, row.text)
    item_id = int(existing_item["id"])
    changed = (
        str(existing_item["text"]) != row.text
        or int(existing_item["lexeme_id"]) != lexeme_id
        or bool(int(existing_item["is_archived"])) != row.is_archived
    )
    if changed:
        connection.execute(
            """
            UPDATE learning_items
            SET lexeme_id = ?,
                text = ?,
                is_archived = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                lexeme_id,
                row.text,
                1 if row.is_archived else 0,
                utc_now(),
                item_id,
            ),
        )
    translation_changed = _sync_translations(connection, item_id, row.translations)
    asset_changed = _sync_media_assets(
        connection,
        item_id,
        row,
        finalized_asset_paths=finalized_asset_paths,
    )
    if changed or translation_changed or asset_changed:
        return item_id, "updated"
    return item_id, "unchanged"


def _create_learning_item(connection: sqlite3.Connection, family_id: int, row: PreparedWorkbookImportRow) -> int:
    cursor = connection.execute(
        """
        INSERT INTO learning_items (
            family_id,
            lexeme_id,
            text,
            is_archived,
            created_at,
            updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            family_id,
            _get_or_create_lexeme_id(connection, row.text),
            row.text,
            1 if row.is_archived else 0,
            utc_now(),
            utc_now(),
        ),
    )
    return int(cursor.lastrowid)


def _sync_translations(
    connection: sqlite3.Connection,
    learning_item_id: int,
    translations: dict[str, str],
) -> bool:
    existing_rows = connection.execute(
        """
        SELECT id, language_code, translation_text
        FROM learning_item_translations
        WHERE learning_item_id = ?
        ORDER BY id
        """,
        (learning_item_id,),
    ).fetchall()
    existing = {str(row["language_code"]): row for row in existing_rows}
    changed = False
    for language_code in ("ru", "uk", "bg"):
        desired_value = translations.get(language_code, "")
        current_row = existing.get(language_code)
        if desired_value:
            if current_row is None:
                connection.execute(
                    """
                    INSERT INTO learning_item_translations (
                        learning_item_id,
                        language_code,
                        translation_text,
                        created_at,
                        updated_at
                    )
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        learning_item_id,
                        language_code,
                        desired_value,
                        utc_now(),
                        utc_now(),
                    ),
                )
                changed = True
            elif str(current_row["translation_text"]) != desired_value:
                connection.execute(
                    """
                    UPDATE learning_item_translations
                    SET translation_text = ?,
                        updated_at = ?
                    WHERE id = ?
                    """,
                    (desired_value, utc_now(), int(current_row["id"])),
                )
                changed = True
        elif current_row is not None:
            connection.execute(
                """
                DELETE FROM learning_item_translations
                WHERE id = ?
                """,
                (int(current_row["id"]),),
            )
            changed = True
    return changed


def _sync_media_assets(
    connection: sqlite3.Connection,
    learning_item_id: int,
    row: PreparedWorkbookImportRow,
    *,
    finalized_asset_paths: list[str],
) -> bool:
    image_changed = _replace_asset_ref(
        connection,
        learning_item_id,
        role=PRIMARY_IMAGE_ROLE,
        asset_type=ASSET_TYPE_IMAGE,
        prepared_asset=row.image_asset,
        finalized_asset_paths=finalized_asset_paths,
    )
    audio_changed = _replace_asset_ref(
        connection,
        learning_item_id,
        role=PRIMARY_AUDIO_ROLE,
        asset_type=ASSET_TYPE_AUDIO,
        prepared_asset=row.audio_asset,
        finalized_asset_paths=finalized_asset_paths,
    )
    return image_changed or audio_changed


def _replace_asset_ref(
    connection: sqlite3.Connection,
    learning_item_id: int,
    *,
    role: str,
    asset_type: str,
    prepared_asset: PreparedAssetRef | None,
    finalized_asset_paths: list[str],
) -> bool:
    existing_assets = connection.execute(
        """
        SELECT assets.local_path, assets.source_url
        FROM learning_item_assets
        JOIN assets
          ON assets.id = learning_item_assets.asset_id
        WHERE learning_item_assets.learning_item_id = ?
          AND learning_item_assets.role = ?
        ORDER BY learning_item_assets.sort_order, learning_item_assets.id
        """,
        (learning_item_id, role),
    ).fetchall()
    current_local_path = ""
    current_source_url = ""
    if existing_assets:
        current_local_path = str(existing_assets[0]["local_path"] or "")
        current_source_url = str(existing_assets[0]["source_url"] or "")

    if prepared_asset is None:
        if not current_local_path and not current_source_url:
            return False
        replace_learning_item_assets_for_role(
            learning_item_id,
            role,
            assets=[],
            connection=connection,
        )
        return True

    desired_source_url = prepared_asset.source_url
    desired_local_path = prepared_asset.local_path
    if current_local_path == desired_local_path and current_source_url == (desired_source_url or ""):
        return False

    if prepared_asset.staged_local_path:
        cursor = connection.execute(
            """
            INSERT INTO assets (
                workbook_key,
                asset_type,
                source_url,
                local_path,
                created_at
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                None,
                asset_type,
                desired_source_url,
                prepared_asset.staged_local_path,
                utc_now(),
            ),
        )
        asset_id = int(cursor.lastrowid)
        desired_local_path = finalize_workbook_import_asset(
            asset_id,
            asset_type,
            prepared_asset.staged_local_path,
        )
        connection.execute(
            """
            UPDATE assets
            SET local_path = ?
            WHERE id = ?
            """,
            (desired_local_path, asset_id),
        )
        finalized_asset_paths.append(desired_local_path)
        replace_learning_item_assets_for_role(
            learning_item_id,
            role,
            assets=[{"asset_id": asset_id}],
            connection=connection,
        )
        return True

    replace_learning_item_assets_for_role(
        learning_item_id,
        role,
        assets=[{
            "asset_type": asset_type,
            "source_url": desired_source_url,
            "local_path": desired_local_path,
        }],
        connection=connection,
    )
    return True


def _sync_topics(
    connection: sqlite3.Connection,
    family_id: int,
    existing_topics: dict[str, sqlite3.Row],
    active_topic_membership: dict[str, list[int]],
) -> None:
    active_topic_identities = {_normalize_topic_identity(title) for title in active_topic_membership}
    for topic_title, learning_item_ids in active_topic_membership.items():
        identity = _normalize_topic_identity(topic_title)
        existing_topic = existing_topics.get(identity)
        if existing_topic is None:
            cursor = connection.execute(
                """
                INSERT INTO topics (
                    family_id,
                    name,
                    title,
                    is_archived,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, 0, ?, ?)
                """,
                (
                    family_id,
                    _build_topic_name(identity),
                    topic_title,
                    utc_now(),
                    utc_now(),
                ),
            )
            topic_id = int(cursor.lastrowid)
        else:
            topic_id = int(existing_topic["id"])
            connection.execute(
                """
                UPDATE topics
                SET name = ?,
                    title = ?,
                    is_archived = 0,
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    _build_topic_name(identity),
                    topic_title,
                    utc_now(),
                    topic_id,
                ),
            )
        connection.execute(
            """
            DELETE FROM topic_items
            WHERE topic_id = ?
            """,
            (topic_id,),
        )
        for learning_item_id in learning_item_ids:
            connection.execute(
                """
                INSERT INTO topic_items (
                    topic_id,
                    learning_item_id
                )
                VALUES (?, ?)
                """,
                (topic_id, learning_item_id),
            )

    for identity, topic in existing_topics.items():
        if identity in active_topic_identities:
            continue
        connection.execute(
            """
            UPDATE topics
            SET is_archived = 1,
                updated_at = ?
            WHERE id = ?
            """,
            (utc_now(), int(topic["id"])),
        )


def _archive_learning_item(connection: sqlite3.Connection, learning_item_id: int) -> None:
    connection.execute(
        """
        UPDATE learning_items
        SET is_archived = 1,
            updated_at = ?
        WHERE id = ?
        """,
        (utc_now(), learning_item_id),
    )


def _get_or_create_lexeme_id(connection: sqlite3.Connection, lemma: str) -> int:
    existing = connection.execute(
        """
        SELECT id
        FROM lexemes
        WHERE lemma = ?
        LIMIT 1
        """,
        (lemma,),
    ).fetchone()
    if existing is not None:
        return int(existing["id"])
    cursor = connection.execute(
        """
        INSERT INTO lexemes (lemma, created_at, updated_at)
        VALUES (?, ?, ?)
        """,
        (lemma, utc_now(), utc_now()),
    )
    return int(cursor.lastrowid)


def _cleanup_staged_asset_paths(asset_paths: tuple[str, ...] | list[str]) -> None:
    for asset_path in asset_paths:
        try:
            resolve_runtime_asset_path(asset_path).unlink(missing_ok=True)
        except OSError:
            continue


def _report_progress(
    progress_callback: ProgressCallback | None,
    event: WorkbookImportProgress,
) -> None:
    if progress_callback is None:
        return
    parameter_count = len(signature(progress_callback).parameters)
    if parameter_count <= 1:
        progress_callback(event)
        return
    progress_callback(event.processed_rows, event.total_rows)


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_item_key(value: str) -> int | None:
    if not value:
        return None
    match = re.fullmatch(r"item-(\d+)", value.strip())
    if match is None:
        return None
    return int(match.group(1))


def _parse_boolish(value: str) -> bool:
    normalized = value.strip().lower()
    if normalized in {"", "0", "false", "no"}:
        return False
    if normalized in {"1", "true", "yes"}:
        return True
    try:
        numeric_value = float(normalized)
    except ValueError:
        numeric_value = None
    if numeric_value == 0.0:
        return False
    if numeric_value == 1.0:
        return True
    raise ValueError("invalid boolean value")


def _parse_topic_titles(value: str) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for raw_topic in re.split(r"[\n,;]+", value):
        topic_title = raw_topic.strip()
        if not topic_title:
            continue
        identity = _normalize_topic_identity(topic_title)
        if identity in seen:
            continue
        seen.add(identity)
        result.append(topic_title)
    return result


def _normalize_topic_identity(topic_title: str) -> str:
    return re.sub(r"\s+", " ", topic_title.strip().casefold())


def _build_topic_name(identity: str) -> str:
    base_name = TOPIC_NAME_PATTERN.sub("-", identity).strip("-")
    return base_name or "topic"
