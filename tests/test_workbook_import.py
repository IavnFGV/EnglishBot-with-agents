import sys
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from PIL import Image

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from englishbot import db
from englishbot.assets import PRIMARY_AUDIO_ROLE, PRIMARY_IMAGE_ROLE, resolve_asset_ref_for_role, resolve_runtime_asset_path
from englishbot.families import create_family, create_family_learning_item, create_family_topic, list_family_topics, replace_topic_items
from englishbot.topics import get_topic_learning_item_ids
from englishbot.vocabulary import get_learning_item, list_learning_item_translations, create_learning_item_translation, create_lexeme
from englishbot.workbook_export import export_family_workbook
from englishbot.workbook_import import (
    WorkbookImportValidationError,
    apply_family_workbook_import,
    apply_prepared_family_workbook_import,
    prepare_family_workbook_import,
    validate_family_workbook_import,
)


def setup_db(tmp_path: Path, monkeypatch) -> int:
    db.DB_PATH = tmp_path / "workbook_import.sqlite3"
    monkeypatch.setenv("ENGLISHBOT_BACKUP_DIR", str(tmp_path / "backups"))
    db.init_db()
    owner = type("User", (), {"id": 1401, "username": "owner", "first_name": "Owner", "last_name": None})()
    db.save_user(owner)
    family = create_family("Home", owner.id)
    return int(family["id"])


def seed_family_content(family_id: int) -> tuple[int, int]:
    first_item_id = create_family_learning_item(family_id, create_lexeme("apple"), "apple")
    second_item_id = create_family_learning_item(family_id, create_lexeme("pear"), "pear")
    create_learning_item_translation(first_item_id, "ru", "яблоко")
    create_learning_item_translation(second_item_id, "ru", "груша")
    topic_id = create_family_topic(family_id, "fruits", "Fruits")
    replace_topic_items(topic_id, [first_item_id, second_item_id])
    with db.get_connection() as connection:
        from englishbot.assets import replace_learning_item_assets_for_role, ASSET_TYPE_IMAGE

        replace_learning_item_assets_for_role(
            first_item_id,
            PRIMARY_IMAGE_ROLE,
            assets=[{"asset_type": ASSET_TYPE_IMAGE, "local_path": "assets/images/apple.jpg"}],
            connection=connection,
        )
    return first_item_id, second_item_id


def make_png_bytes() -> bytes:
    buffer = BytesIO()
    image = Image.new("RGB", (1, 1), color="red")
    image.save(buffer, format="PNG")
    return buffer.getvalue()


def make_mp3_like_bytes() -> bytes:
    return b"ID3" + b"\x00" * 32


def test_import_updates_creates_archives_and_ignores_display_only_image_column(tmp_path: Path, monkeypatch) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, second_item_id = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family.xlsx").file_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["learning_items"]

    sheet["B2"] = "green apple"
    sheet["C2"] = "зеленое яблоко"
    sheet["F2"] = "Fruits\nFresh picks"
    sheet["H2"] = '=IMAGE("https://malicious.example.com/override.jpg")'
    sheet.delete_rows(3, 1)
    sheet.append(
        [
            "",
            "plum",
            "слива",
            "",
            "",
            "Fresh picks",
            "assets/images/plum.jpg",
            '=IMAGE("https://example.com/plum.jpg")',
            "",
            0,
        ]
    )
    workbook.save(workbook_path)

    summary = apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    updated_item = get_learning_item(first_item_id)
    archived_item = get_learning_item(second_item_id)
    all_topics = list_family_topics(family_id)
    fresh_topic = next(topic for topic in all_topics if topic["title"] == "Fresh picks")
    translations = {
        row["language_code"]: row["translation_text"]
        for row in list_learning_item_translations(first_item_id)
    }
    family_items = {
        item["text"]: item
        for item in [
            get_learning_item(first_item_id),
            get_learning_item(second_item_id),
            get_learning_item(max(first_item_id, second_item_id) + 1),
        ]
        if item is not None and item["family_id"] == family_id
    }

    assert summary.created == 1
    assert summary.updated == 1
    assert summary.archived == 1
    assert summary.backup_file_path.exists()
    assert updated_item is not None
    assert updated_item["text"] == "green apple"
    assert translations["ru"] == "зеленое яблоко"
    assert resolve_asset_ref_for_role(first_item_id, PRIMARY_IMAGE_ROLE) == "assets/images/apple.jpg"
    assert archived_item is not None
    assert archived_item["is_archived"] == 1
    assert "plum" in family_items
    assert get_topic_learning_item_ids(int(fresh_topic["id"])) == [first_item_id, int(family_items["plum"]["id"])]


def test_import_validates_before_writing(tmp_path: Path, monkeypatch) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, _ = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-invalid.xlsx").file_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["learning_items"]
    sheet.append(list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0])
    workbook.save(workbook_path)

    with pytest.raises(WorkbookImportValidationError):
        apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    assert get_learning_item(first_item_id)["text"] == "apple"


def test_import_runs_in_one_transaction(tmp_path: Path, monkeypatch) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, _ = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-atomic.xlsx").file_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["learning_items"]
    sheet["B2"] = "changed apple"
    workbook.save(workbook_path)

    def fail_replace_asset_ref(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr("englishbot.workbook_import._replace_asset_ref", fail_replace_asset_ref)

    with pytest.raises(RuntimeError):
        apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    assert get_learning_item(first_item_id)["text"] == "apple"


def test_import_accepts_excel_style_float_archive_flags(tmp_path: Path, monkeypatch) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, _ = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-floats.xlsx").file_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["learning_items"]
    sheet["J2"] = 0.0
    workbook.save(workbook_path)

    summary = apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    assert summary.updated == 0
    assert get_learning_item(first_item_id)["is_archived"] == 0


def test_import_reports_row_level_progress(tmp_path: Path, monkeypatch) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-progress.xlsx").file_path
    progress_events: list[tuple[str, int, int]] = []

    summary = apply_family_workbook_import(
        workbook_path,
        family_id,
        started_by_user_id=1401,
        progress_callback=lambda event: progress_events.append(
            (event.phase, event.processed_rows, event.total_rows)
        ),
    )

    assert summary.unchanged == 2
    assert progress_events == [
        ("preparing", 1, 2),
        ("preparing", 2, 2),
        ("applying", 1, 2),
        ("applying", 2, 2),
    ]


def test_import_accepts_exported_public_static_asset_urls_without_network(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv(
        "INFRA_STATIC_BASE_URL",
        "https://englishbot-178-104-84-123.nip.io/static/hwgpLEf0YVr_FyXgh2ZFSyj-EMwN6IQ8m2leJ6XRS1k",
    )
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, second_item_id = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-static-roundtrip.xlsx").file_path

    def fail_network(*args, **kwargs):
        raise AssertionError("network should not be used for exported static asset URLs")

    monkeypatch.setattr("englishbot.workbook_import.download_remote_asset_content", fail_network)

    summary = apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    assert summary.unchanged == 2
    assert resolve_asset_ref_for_role(first_item_id, PRIMARY_IMAGE_ROLE) == "assets/images/apple.jpg"
    assert get_learning_item(second_item_id)["text"] == "pear"


def test_backup_failure_prevents_any_import_work_from_starting(tmp_path: Path, monkeypatch) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, _ = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-backup-fail.xlsx").file_path
    prepare_called = False

    def fail_backup(*args, **kwargs):
        raise RuntimeError("backup failed")

    def fake_prepare(*args, **kwargs):
        nonlocal prepare_called
        prepare_called = True
        return SimpleNamespace(rows=[], staged_asset_paths=())

    monkeypatch.setattr("englishbot.workbook_import.create_bulk_edit_backup", fail_backup)
    monkeypatch.setattr("englishbot.workbook_import.prepare_family_workbook_import", fake_prepare)

    with pytest.raises(RuntimeError, match="backup failed"):
        apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    assert prepare_called is False
    assert get_learning_item(first_item_id)["text"] == "apple"


def test_failed_remote_asset_download_leaves_sqlite_unchanged_and_cleans_staging(
    tmp_path: Path,
    monkeypatch,
) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, second_item_id = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-remote-fail.xlsx").file_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["learning_items"]
    sheet["G2"] = "https://example.com/apple.png"
    sheet["G3"] = "https://example.com/pear.png"
    workbook.save(workbook_path)

    calls = {"count": 0}

    def fake_download(*args, **kwargs):
        calls["count"] += 1
        if calls["count"] == 1:
            return make_png_bytes()
        raise RuntimeError("download failed")

    monkeypatch.setattr("englishbot.workbook_import.download_remote_asset_content", fake_download)

    with pytest.raises(RuntimeError, match="download failed"):
        apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    assert get_learning_item(first_item_id)["text"] == "apple"
    assert get_learning_item(second_item_id)["text"] == "pear"
    staged_dir = resolve_runtime_asset_path("assets/import-staging/image")
    assert not staged_dir.exists() or list(staged_dir.iterdir()) == []


def test_failed_image_validation_leaves_sqlite_unchanged(tmp_path: Path, monkeypatch) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, _ = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-invalid-image.xlsx").file_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["learning_items"]
    sheet["G2"] = "https://example.com/apple.png"
    workbook.save(workbook_path)

    def fail_validation(*args, **kwargs):
        raise ValueError("image content from https://example.com/apple.png is not a valid image")

    monkeypatch.setattr("englishbot.workbook_import.download_remote_asset_content", fail_validation)

    with pytest.raises(ValueError, match="not a valid image"):
        apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    assert get_learning_item(first_item_id)["text"] == "apple"


def test_apply_phase_uses_prepared_local_asset_refs_without_network(tmp_path: Path, monkeypatch) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, _ = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-prepare-apply.xlsx").file_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["learning_items"]
    sheet["G2"] = "https://example.com/apple.png"
    workbook.save(workbook_path)

    monkeypatch.setattr(
        "englishbot.workbook_import.download_remote_asset_content",
        lambda *args, **kwargs: make_png_bytes(),
    )
    validated_rows = validate_family_workbook_import(workbook_path, family_id)
    prepared_import = prepare_family_workbook_import(workbook_path, family_id, validated_rows)

    def fail_network(*args, **kwargs):
        raise AssertionError("network should not be used during apply")

    monkeypatch.setattr("englishbot.workbook_import.download_remote_asset_content", fail_network)

    summary = apply_prepared_family_workbook_import(
        family_id,
        prepared_import,
        tmp_path / "backups" / "prepared.sqlite3",
    )

    assert summary.updated == 1
    assert resolve_asset_ref_for_role(first_item_id, PRIMARY_IMAGE_ROLE).startswith("assets/images/imported/")


def test_remote_image_import_uses_persistent_imported_path_and_asset_id_filename(
    tmp_path: Path,
    monkeypatch,
) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, _ = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-imported-image.xlsx").file_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["learning_items"]
    sheet["G2"] = "https://example.com/image-without-extension"
    workbook.save(workbook_path)

    monkeypatch.setattr(
        "englishbot.workbook_import.download_remote_asset_content",
        lambda *args, **kwargs: make_png_bytes(),
    )

    summary = apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    image_ref = resolve_asset_ref_for_role(first_item_id, PRIMARY_IMAGE_ROLE)
    assert summary.updated == 1
    assert image_ref is not None
    assert image_ref.startswith("assets/images/imported/")
    assert image_ref.endswith(".png")
    with db.get_connection() as connection:
        asset_row = connection.execute(
            """
            SELECT assets.id, assets.local_path, assets.source_url
            FROM learning_item_assets
            JOIN assets ON assets.id = learning_item_assets.asset_id
            WHERE learning_item_assets.learning_item_id = ?
              AND learning_item_assets.role = ?
            """,
            (first_item_id, PRIMARY_IMAGE_ROLE),
        ).fetchone()
    assert asset_row is not None
    assert asset_row["source_url"] == "https://example.com/image-without-extension"
    assert asset_row["local_path"] == f"assets/images/imported/asset-{int(asset_row['id'])}.png"
    assert resolve_runtime_asset_path(str(asset_row["local_path"])).exists()
    image_staging_dir = resolve_runtime_asset_path("assets/import-staging/image")
    assert not image_staging_dir.exists() or list(image_staging_dir.iterdir()) == []


def test_remote_audio_import_uses_persistent_imported_path_and_sensible_extension(
    tmp_path: Path,
    monkeypatch,
) -> None:
    family_id = setup_db(tmp_path, monkeypatch)
    first_item_id, _ = seed_family_content(family_id)
    workbook_path = export_family_workbook(family_id, output_path=tmp_path / "family-imported-audio.xlsx").file_path
    workbook = load_workbook(workbook_path)
    sheet = workbook["learning_items"]
    sheet["I2"] = "https://example.com/audio-without-extension"
    workbook.save(workbook_path)

    monkeypatch.setattr(
        "englishbot.workbook_import.download_remote_asset_content",
        lambda *args, **kwargs: make_mp3_like_bytes(),
    )

    summary = apply_family_workbook_import(workbook_path, family_id, started_by_user_id=1401)

    audio_ref = resolve_asset_ref_for_role(first_item_id, PRIMARY_AUDIO_ROLE)
    assert summary.updated == 1
    assert audio_ref is not None
    assert audio_ref.startswith("assets/audio/imported/")
    assert audio_ref.endswith(".mp3")
    assert not audio_ref.endswith(".bin")
    with db.get_connection() as connection:
        asset_row = connection.execute(
            """
            SELECT assets.id, assets.local_path, assets.source_url
            FROM learning_item_assets
            JOIN assets ON assets.id = learning_item_assets.asset_id
            WHERE learning_item_assets.learning_item_id = ?
              AND learning_item_assets.role = ?
            """,
            (first_item_id, PRIMARY_AUDIO_ROLE),
        ).fetchone()
    assert asset_row is not None
    assert asset_row["source_url"] == "https://example.com/audio-without-extension"
    assert asset_row["local_path"] == f"assets/audio/imported/asset-{int(asset_row['id'])}.mp3"
    assert resolve_runtime_asset_path(str(asset_row["local_path"])).exists()
    audio_staging_dir = resolve_runtime_asset_path("assets/import-staging/audio")
    assert not audio_staging_dir.exists() or list(audio_staging_dir.iterdir()) == []
