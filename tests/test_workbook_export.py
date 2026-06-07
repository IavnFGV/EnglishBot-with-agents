import sys
from pathlib import Path

from openpyxl import load_workbook
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from englishbot import db
from englishbot.assets import ASSET_TYPE_IMAGE, PRIMARY_IMAGE_ROLE, create_asset, link_asset_to_learning_item
from englishbot.families import create_family, create_family_learning_item, create_family_topic, replace_topic_items
from englishbot.workbook_export import (
    IMAGE_COLUMN_LETTER,
    IMAGE_COLUMN_WIDTH,
    LEARNING_ITEM_ROW_HEIGHT_POINTS,
    export_family_workbook,
)
from englishbot.vocabulary import create_learning_item_translation, create_lexeme


def setup_db(tmp_path: Path) -> None:
    db.DB_PATH = tmp_path / "workbook_export.sqlite3"
    db.init_db()


def test_export_creates_valid_xlsx_with_display_only_image_formula(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.delenv("INFRA_STATIC_BASE_URL", raising=False)
    setup_db(tmp_path)
    owner = type("User", (), {"id": 1301, "username": "owner", "first_name": "Owner", "last_name": None})()
    db.save_user(owner)
    family = create_family("Home", owner.id)
    family_id = int(family["id"])
    item_id = create_family_learning_item(family_id, create_lexeme("apple"), "apple")
    create_learning_item_translation(item_id, "ru", "яблоко")
    topic_id = create_family_topic(family_id, "fruits", "Fruits")
    replace_topic_items(topic_id, [item_id])
    asset_id = create_asset(
        ASSET_TYPE_IMAGE,
        source_url="https://example.com/apple.jpg",
        local_path="assets/images/apple.jpg",
    )
    link_asset_to_learning_item(item_id, asset_id, PRIMARY_IMAGE_ROLE)

    result = export_family_workbook(family_id, output_path=tmp_path / "family.xlsx")
    workbook = load_workbook(result.file_path, data_only=False)
    sheet = workbook["learning_items"]

    assert workbook.sheetnames == ["meta", "learning_items"]
    assert result.topic_count == 1
    assert result.learning_item_count == 1
    assert [cell.value for cell in sheet[1]] == [
        "item_key",
        "text",
        "translation_ru",
        "translation_uk",
        "translation_bg",
        "topics",
        "image_ref",
        "image",
        "audio_ref",
        "is_archived",
    ]
    assert sheet["G2"].value == "https://example.com/apple.jpg"
    assert sheet["H2"].value == "=IMAGE(G2)"
    assert sheet.column_dimensions[IMAGE_COLUMN_LETTER].width == IMAGE_COLUMN_WIDTH
    assert sheet.row_dimensions[2].height == LEARNING_ITEM_ROW_HEIGHT_POINTS


def test_export_uses_public_static_url_for_local_image_refs(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv(
        "INFRA_STATIC_BASE_URL",
        "https://englishbot-178-104-84-123.nip.io/static/hwgpLEf0YVr_FyXgh2ZFSyj-EMwN6IQ8m2leJ6XRS1k",
    )
    setup_db(tmp_path)
    owner = type("User", (), {"id": 1302, "username": "owner", "first_name": "Owner", "last_name": None})()
    db.save_user(owner)
    family = create_family("Home", owner.id)
    family_id = int(family["id"])
    item_id = create_family_learning_item(family_id, create_lexeme("apple"), "apple")
    asset_id = create_asset(
        ASSET_TYPE_IMAGE,
        source_url="https://example.com/apple.jpg",
        local_path="assets/images/apple.jpg",
    )
    link_asset_to_learning_item(item_id, asset_id, PRIMARY_IMAGE_ROLE)

    result = export_family_workbook(family_id, output_path=tmp_path / "family.xlsx")
    workbook = load_workbook(result.file_path, data_only=False)
    sheet = workbook["learning_items"]

    assert sheet["G2"].value == (
        "https://englishbot-178-104-84-123.nip.io/static/"
        "hwgpLEf0YVr_FyXgh2ZFSyj-EMwN6IQ8m2leJ6XRS1k/images/apple.jpg"
    )
    assert sheet["H2"].value == "=IMAGE(G2)"
    assert sheet.column_dimensions[IMAGE_COLUMN_LETTER].width == IMAGE_COLUMN_WIDTH
    assert sheet.row_dimensions[2].height == LEARNING_ITEM_ROW_HEIGHT_POINTS
